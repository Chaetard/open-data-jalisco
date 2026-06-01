"""XLSX text extractor backed by openpyxl in read-only / data-only mode.

Streams rows lazily to avoid loading the entire workbook into memory. Emits one
``ExtractedPage`` per non-empty worksheet (page numbers follow workbook order).
Worksheets without any non-empty cell are skipped silently — if the whole
workbook has no text, the extractor returns an empty ``ExtractedDocument``
(``needs_ocr=False``) so the pipeline marks the doc as ``no_text``, not
``failed``.
"""
from __future__ import annotations

import datetime as _dt
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from ...ports.text_extractor import ExtractedDocument, ExtractedPage
from ...shared.logging import get_logger

logger = get_logger(__name__)

_XLSX_MIME = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel.sheet.macroEnabled.12",
}
_XLSX_EXTENSIONS = {"xlsx", "xlsm"}

# Cap the number of cells we'll read per worksheet to keep memory predictable
# on accidentally huge files (e.g. someone uploads a 1M-row export).
_MAX_CELLS_PER_SHEET = 200_000


class XlsxTextExtractor:
    def can_handle(self, mime_type: str, extension: str) -> bool:
        return (
            mime_type.lower() in _XLSX_MIME
            or extension.lower().lstrip(".") in _XLSX_EXTENSIONS
        )

    def extract(self, path: Path) -> ExtractedDocument:
        try:
            wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        except (
            InvalidFileException,
            zipfile.BadZipFile,
            OSError,
            KeyError,
            ValueError,
        ) as e:
            # BadZipFile / KeyError / ValueError happen when the file is a
            # renamed non-xlsx or otherwise corrupt. Treat as "no text" rather
            # than crashing the processing batch.
            logger.warning("xlsx.extract.open_error path=%s err=%r", path, e)
            return ExtractedDocument(
                full_text="",
                pages=[],
                needs_ocr=False,
                metadata={"error": f"open_error: {e!r}"},
            )

        try:
            pages: list[ExtractedPage] = []
            sheet_names: list[str] = []
            total_rows = 0
            truncated_sheets: list[str] = []

            for sheet_index, sheet in enumerate(wb.worksheets, start=1):
                sheet_names.append(sheet.title)
                rows_text, row_count, truncated = self._sheet_text(sheet)
                total_rows += row_count
                if truncated:
                    truncated_sheets.append(sheet.title)
                if not rows_text:
                    continue
                header = f"# {sheet.title}\n"
                page_text = header + "\n".join(rows_text)
                pages.append(ExtractedPage(page_number=sheet_index, text=page_text))
        finally:
            # read-only workbooks hold a zipfile handle; close to release it.
            wb.close()

        full_text = "\n\n".join(p.text for p in pages if p.text)

        metadata: dict[str, Any] = {
            "sheet_names": sheet_names,
            "sheet_count": len(sheet_names),
            "non_empty_sheet_count": len(pages),
            "row_count": total_rows,
        }
        if truncated_sheets:
            metadata["truncated_sheets"] = truncated_sheets

        return ExtractedDocument(
            full_text=full_text,
            pages=pages,
            needs_ocr=False,
            metadata=metadata,
        )

    @staticmethod
    def _sheet_text(sheet) -> tuple[list[str], int, bool]:
        """Return (non-empty row strings, total non-empty row count, truncated)."""
        rows_text: list[str] = []
        row_count = 0
        cells_seen = 0
        truncated = False
        for row in sheet.iter_rows(values_only=True):
            row_count_candidate = row_count + 1
            cells = [
                _format_cell(v)
                for v in row
                if v is not None and not (isinstance(v, str) and not v.strip())
            ]
            cells_seen += len(row) if row is not None else 0
            if not cells:
                continue
            rows_text.append(" | ".join(cells))
            row_count = row_count_candidate
            if cells_seen >= _MAX_CELLS_PER_SHEET:
                truncated = True
                logger.warning(
                    "xlsx.extract.sheet_truncated sheet=%s cells_seen=%d limit=%d",
                    sheet.title,
                    cells_seen,
                    _MAX_CELLS_PER_SHEET,
                )
                break
        return rows_text, row_count, truncated


def _format_cell(value: Any) -> str:
    """Cell value → compact string. Dates render as ISO 8601, floats trim trailing zeros."""
    if isinstance(value, _dt.datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, _dt.date):
        return value.isoformat()
    if isinstance(value, float):
        # Avoid scientific notation noise for typical sheet values.
        if value.is_integer():
            return str(int(value))
        # Strip trailing zeros while keeping precision.
        text = f"{value:.10g}"
        return text
    return str(value).strip()
