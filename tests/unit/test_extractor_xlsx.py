"""Unit tests for the XLSX text extractor.

XLSX fixtures are written to ``tmp_path`` with openpyxl in write mode; the
extractor then opens them in read-only/data-only mode (the real path used by
the pipeline). No network, no DB.
"""
from __future__ import annotations

import datetime as _dt
import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from open_data_jalisco.adapters.extraction.xlsx import XlsxTextExtractor


def _write_xlsx(path: Path, build) -> Path:
    """Run ``build(wb)`` on a fresh Workbook and save to ``path``."""
    wb = Workbook()
    # Workbook() comes with a default sheet; tests may rename or remove it.
    build(wb)
    wb.save(str(path))
    return path


def test_can_handle_by_mime_and_extension():
    e = XlsxTextExtractor()
    assert e.can_handle(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"
    )
    assert e.can_handle("application/octet-stream", ".xlsx")
    assert e.can_handle("application/octet-stream", "XLSX")  # case-insensitive
    assert e.can_handle("application/octet-stream", "xlsm")
    assert not e.can_handle("application/pdf", "pdf")
    assert not e.can_handle("text/plain", "txt")


def test_extracts_text_per_sheet(tmp_path: Path):
    def build(wb: Workbook) -> None:
        ws = wb.active
        ws.title = "Sheet One"
        ws.append(["Col A", "Col B", "Col C"])
        ws.append(["alpha", "beta", "gamma"])
        ws.append([None, None, None])           # fully empty row → skipped
        ws.append(["", "  ", ""])               # whitespace-only → skipped
        ws.append(["last", "", "row"])          # partial → kept

        ws2 = wb.create_sheet("Sheet Two")
        ws2.append(["only", "two"])

    path = _write_xlsx(tmp_path / "doc.xlsx", build)

    result = XlsxTextExtractor().extract(path)

    assert not result.needs_ocr
    assert len(result.pages) == 2
    assert result.pages[0].page_number == 1
    assert result.pages[1].page_number == 2
    assert "alpha | beta | gamma" in result.pages[0].text
    assert "last | row" in result.pages[0].text
    # Sheet title shows up as a heading the chunker can use.
    assert result.pages[0].text.startswith("# Sheet One")
    assert result.pages[1].text.startswith("# Sheet Two")
    assert "only | two" in result.pages[1].text
    assert result.metadata["sheet_names"] == ["Sheet One", "Sheet Two"]
    assert result.metadata["sheet_count"] == 2
    assert result.metadata["non_empty_sheet_count"] == 2


def test_completely_empty_workbook_returns_no_pages_not_ocr(tmp_path: Path):
    """An empty XLSX must NOT crash and must NOT trigger needs_ocr."""
    def build(wb: Workbook) -> None:
        wb.active.title = "Blank"   # no rows added

    path = _write_xlsx(tmp_path / "empty.xlsx", build)

    result = XlsxTextExtractor().extract(path)

    assert result.needs_ocr is False
    assert result.pages == []
    assert result.full_text == ""
    assert result.metadata["sheet_count"] == 1
    assert result.metadata["non_empty_sheet_count"] == 0


def test_sheets_without_text_are_skipped_but_others_kept(tmp_path: Path):
    def build(wb: Workbook) -> None:
        wb.active.title = "Empty"
        ws2 = wb.create_sheet("HasData")
        ws2.append(["x", "y"])

    path = _write_xlsx(tmp_path / "mixed.xlsx", build)
    result = XlsxTextExtractor().extract(path)

    assert len(result.pages) == 1
    assert result.pages[0].page_number == 2
    assert "HasData" in result.pages[0].text
    assert result.metadata["sheet_names"] == ["Empty", "HasData"]
    assert result.metadata["non_empty_sheet_count"] == 1


def test_numeric_and_date_cells_render_cleanly(tmp_path: Path):
    def build(wb: Workbook) -> None:
        ws = wb.active
        ws.title = "Types"
        ws.append([1, 2.0, 3.14])
        ws.append([_dt.date(2025, 10, 5), _dt.datetime(2025, 10, 5, 14, 30)])

    path = _write_xlsx(tmp_path / "types.xlsx", build)
    text = XlsxTextExtractor().extract(path).pages[0].text

    # int stays int; float that's actually int renders as "2"; pi keeps precision
    assert "1 | 2 | 3.14" in text
    assert "2025-10-05" in text
    assert "2025-10-05 14:30:00" in text


def test_invalid_file_returns_empty_not_raises(tmp_path: Path):
    bogus = tmp_path / "not-really-xlsx.xlsx"
    bogus.write_bytes(b"this is not a zip file")

    result = XlsxTextExtractor().extract(bogus)

    assert result.full_text == ""
    assert result.pages == []
    assert result.needs_ocr is False
    assert "error" in result.metadata


def test_truncated_zip_xlsx_returns_empty(tmp_path: Path):
    # A valid zip but missing the xlsx structure.
    bogus = tmp_path / "broken.xlsx"
    with zipfile.ZipFile(bogus, "w") as zf:
        zf.writestr("hello.txt", "not an xlsx")

    result = XlsxTextExtractor().extract(bogus)

    assert result.pages == []
    assert result.needs_ocr is False
    assert "error" in result.metadata


def test_extracted_full_text_concatenates_pages(tmp_path: Path):
    def build(wb: Workbook) -> None:
        ws = wb.active
        ws.title = "A"
        ws.append(["alpha"])
        ws2 = wb.create_sheet("B")
        ws2.append(["bravo"])

    path = _write_xlsx(tmp_path / "concat.xlsx", build)
    result = XlsxTextExtractor().extract(path)

    assert "alpha" in result.full_text
    assert "bravo" in result.full_text
    # Pages are joined with a blank line.
    assert result.full_text == "\n\n".join(p.text for p in result.pages)


@pytest.mark.parametrize("ext", ["xlsx", "xlsm", ".xlsx", "XLSX"])
def test_extension_is_case_and_dot_insensitive(ext: str):
    assert XlsxTextExtractor().can_handle("application/octet-stream", ext)
